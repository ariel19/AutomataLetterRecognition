#!/usr/bin/env python

import random
import openpyxl

filename = 'input.dat'

class_num = 0
feat_num = 0
min_los = []
max_los = []

nondeterministic_proportions = 0.0

################

is_foreign = False

test_data_from_file = False
foreign_data_from_prop = True
foreign_test_from_file = False

splits_num = 0

train_file = ''
test_file = ''
foreign_train_file = ''
foreign_test_file = ''

test_proportions = 0.0
foreign_proportions = 0.0

pso_args = ()

letters = []
test_set = []
train_set = []
foreign_test_set = []
foreign_train_set = []

train_size = 0
test_size = 0
foreign_train_size = 0
foreign_test_size = 0


def read_train_test(path, tab):
    wb = openpyxl.load_workbook(path, True)
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    base_letter_id = 1 if is_foreign else 1

    for row in ws.iter_rows():
        letter = []
        letter_id = base_letter_id
        first = True
        for cell in row:
            if first:
                first = False
                letter_id += int(cell.internal_value)
            else:
                letter.append(float(cell.internal_value))
        tab.append((letter, letter_id))


def create_letters_sets():
    for i in range(0, train_size):
        n = random.randint(0, len(letters) - 1)
        train_set.append(letters[n])
        del letters[n]

    for i in range(0, test_size):
        n = random.randint(0, len(letters) - 1)
        test_set.append(letters[n])
        del letters[n]


def create_min_max_los():
    global min_los, max_los
    min_los = [0] * feat_num
    max_los = [0] * feat_num

    first = True

    files = [train_file, test_file, foreign_train_file, foreign_test_file]
    for f in files:
        if f != '':
            wb = openpyxl.load_workbook(f, True)
            ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])

            for row in ws.iter_rows():
                first_cell = True
                ncell = 0
                for cell in row:
                    if first_cell:
                        first_cell = False
                    else:
                        if first:
                            min_los[ncell] = float(cell.internal_value)
                            max_los[ncell] = float(cell.internal_value)
                        if float(cell.internal_value) < min_los[ncell]:
                            min_los[ncell] = float(cell.internal_value)
                        if float(cell.internal_value) > max_los[ncell]:
                            max_los[ncell] = float(cell.internal_value)
                        ncell += 1
                first = False


def gen_dat():
    global train_set, test_set
    global train_size, test_size, foreign_train_size, foreign_test_size
    global feat_num, class_num, max_los, min_los

    read_train_test(train_file, letters)
    feat_num = len(letters[0][0])
    for i in letters:
        if i[1] > class_num:
            class_num = i[1]
    if is_foreign:
        class_num -= 1
    create_min_max_los()

    if test_data_from_file:
        train_set = letters
        read_train_test(test_file, test_set)
        train_size = len(train_set)
        test_size = len(test_set)
    else:
        all_s = len(letters)
        test_size = int(all_s * (test_proportions / 100.0))
        train_size = all_s - test_size
        create_letters_sets()

    if is_foreign:
        if foreign_data_from_prop:
            foreign_train_size = int(train_size * (foreign_proportions / 100.0))
            foreign_test_size = int(test_size * (foreign_proportions / 100.0))

            for i in range(0, foreign_train_size):
                letter = []
                for j in range(0, feat_num):
                    letter.append(random.uniform(min_los[j], max_los[j]))
                foreign_train_set.append((letter, 0))

            for i in range(0, foreign_test_size):
                letter = []
                for j in range(0, feat_num):
                    letter.append(random.uniform(min_los[j], max_los[j]))
                foreign_test_set.append((letter, 0))

        else:
            read_train_test(foreign_train_file, foreign_train_set)
            foreign_train_size = len(foreign_train_set)
            if foreign_test_file != '':
                read_train_test(foreign_test_file, foreign_test_set)
                foreign_test_size = len(foreign_test_set)
            elif not test_data_from_file:
                foreign_test_size = int(foreign_train_size * (test_proportions / 100.0))
                for i in range(0, foreign_test_size):
                    letter = []
                    for j in range(0, feat_num):
                        letter.append(random.uniform(min_los[j], max_los[j]))
                    foreign_test_set.append((letter, 0))

    f = open(filename, 'w')

    f.write('1\n')
    f.write(str(class_num) + ', ' + str(feat_num) + ', ' + str(splits_num) + ', ' +
            str(train_size + foreign_train_size) + ', ' + str(test_size + foreign_test_size) + ', ' +
            str(nondeterministic_proportions) + ', \n')

    for a in pso_args:
        f.write(str(a) + ', ')
    f.write('\n\n')

    for i in min_los:
        f.write(str(i) + ', ')
    f.write('\n')

    for i in max_los:
        f.write((str(i) + ', '))
    f.write('\n\n')

    for l in train_set:
        f.write(str(l[1]) + '\n')
        for ft in l[0]:
            f.write(str(ft) + ', ')
        f.write('\n\n')

    for l in foreign_train_set:
        f.write('0\n')
        for ft in l[0]:
            f.write(str(ft) + ', ')
        f.write('\n\n')

    for l in test_set:
        f.write(str(l[1]) + '\n')
        for ft in l[0]:
            f.write(str(ft) + ', ')
        f.write('\n\n')

    for l in foreign_test_set:
        f.write('0\n')
        for ft in l[0]:
            f.write(str(ft) + ', ')
        f.write('\n\n')

    f.close()


def generate_all_data(_splits, _train_file, _test_param, _is_test_file, _is_for_prop, _for_train_file, _for_test_file, _for_prop, _nondet_prop, _pso_args, _is_for):
    global splits_num, train_file, pso_args, is_foreign
    global test_data_from_file, test_file, test_proportions
    global foreign_data_from_prop, foreign_test_from_file
    global foreign_test_file, foreign_train_file, foreign_proportions
    global nondeterministic_proportions

    splits_num = int(_splits)
    train_file = _train_file
    pso_args = _pso_args

    test_data_from_file = _is_test_file
    if test_data_from_file:
        test_file = _test_param
    else:
        test_proportions = float(_test_param)

    foreign_data_from_prop = _is_for_prop
    foreign_test_file = _for_test_file
    foreign_train_file = _for_train_file
    foreign_proportions = float(_for_prop)

    is_foreign = _is_for
    nondeterministic_proportions = float(_nondet_prop)

    gen_dat()


def generate_data_a1(_splits, _train_file, _test_param, _is_test_file, _pso_args):
    generate_all_data(_splits, _train_file, _test_param, _is_test_file, False, '', '', 0.0, 0.0, _pso_args, False)


def generate_data_a2(_splits, _train_file, _test_param, _is_test_file, _is_for_prop, _for_train_file, _for_test_file, _for_prop, _pso_args):
    generate_all_data(_splits, _train_file, _test_param, _is_test_file, _is_for_prop, _for_train_file,
                      _for_test_file, _for_prop, 0.0, _pso_args, True)


def generate_data_a3(_splits, _train_file, _test_param, _is_test_file, _nondet_prop, _pso_args):
    generate_all_data(_splits, _train_file, _test_param, _is_test_file, False, '', '', 0.0, _nondet_prop, _pso_args, False)


def generate_data_a4(_splits, _train_file, _test_param, _is_test_file, _is_for_prop, _for_train_file, _for_test_file, _for_prop, _nondet_prop, _pso_args):
    generate_all_data(_splits, _train_file, _test_param, _is_test_file, _is_for_prop, _for_train_file,
                      _for_test_file, _for_prop, _nondet_prop, _pso_args, True)


def generate_data_a5(_splits, _train_file, _test_param, _is_test_file, _pso_args):
    generate_all_data(_splits, _train_file, _test_param, _is_test_file, False, '', '', 0.0, 0.0, _pso_args, False)


def generate_data_a6(_splits, _train_file, _test_param, _is_test_file, _is_for_prop, _for_train_file, _for_test_file, _for_prop, _pso_args):
    generate_all_data(_splits, _train_file, _test_param, _is_test_file, _is_for_prop, _for_train_file,
                      _for_test_file, _for_prop, 0.0, _pso_args, True)